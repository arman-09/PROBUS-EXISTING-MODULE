"""
modules/routes.py — All Flask API Endpoints
=============================================
Registered via init(app, ...) called from app.py
"""

import threading, logging, os, json, secrets
from datetime import datetime, timedelta
from flask import request, jsonify, session, redirect, Response
try:
    from werkzeug.security import generate_password_hash, check_password_hash
    _AUTH_LIB_AVAILABLE = True
except ImportError:
    _AUTH_LIB_AVAILABLE = False

log = logging.getLogger("routes")

# Field-level diversion marking — fully separate module, see its own
# docstring for the design rationale. Optional import: if the module is
# missing, the related API routes simply return a clear error instead of
# preventing the rest of routes.py (and the whole app) from loading.
try:
    from modules.field_diversion import (
        init_table as _fd_init_table,
        mark_field_diversion as _fd_mark,
        clear_field_mark as _fd_clear,
        get_active_mark as _fd_get_active,
        get_all_active_marks as _fd_get_all,
        feeder_picker_options as _fd_picker_options,
    )
    _FIELD_DIVERSION_AVAILABLE = True
except ImportError:
    _FIELD_DIVERSION_AVAILABLE = False
    log.warning("modules.field_diversion not found — field-level diversion "
                "marking API routes will return 503")

# Weather integration — same optional/graceful pattern. mbscada_scraper.py
# is the existing STLF module's scraper (provided separately); we only
# import the small helpers needed for the Settings UI (save credentials,
# test connection) — the actual background polling loop is started from
# app.py, not from here.
try:
    import modules.mbscada_scraper as _mb
    _MBSCADA_AVAILABLE = True
except ImportError:
    _MBSCADA_AVAILABLE = False
    log.warning("mbscada_scraper module not found — weather config/test "
                "API routes will return 503")

# Service references (set by init)
_cfg = _scraper = _detector = _fm = _store = _email = _wa = _peak = _mgmt = _notify_queue = None


def init(app, cfg, scraper, detector, feeder_master, alert_store, email_mgr, wa_mgr,
         peak_store=None, mgmt_reporter=None, notify_queue=None):
    global _cfg, _scraper, _detector, _fm, _store, _email, _wa, _peak, _mgmt, _notify_queue
    _cfg = cfg; _scraper = scraper; _detector = detector
    _fm = feeder_master; _store = alert_store
    _email = email_mgr; _wa = wa_mgr; _peak = peak_store; _mgmt = mgmt_reporter
    _notify_queue = notify_queue
    if _FIELD_DIVERSION_AVAILABLE:
        try:
            _fd_init_table()
        except Exception as e:
            log.error(f"field_diversion.init_table() failed: {e}")
    _register(app)


def _register(app):

    # ── Authentication gate ──────────────────────────────────────────────
    # Addresses: dashboard currently has NO login — anyone who knows the
    # IP/port on the network can open it and act on live grid alerts.
    # Session-based: one password for the whole dashboard (matches a
    # small-ops-team tool — not meant to replace per-user accounts, just
    # to stop "anyone on the LAN" from walking straight in).

    # Stable signed-cookie secret across restarts — generated once,
    # persisted next to config.json. Without this, every app restart
    # would invalidate everyone's session (forcing re-login constantly).
    if not getattr(app, "secret_key", None):
        cfg_dir = os.path.dirname(_cfg._path) if _cfg and getattr(_cfg, "_path", None) else "data"
        key_path = os.path.join(cfg_dir or ".", "secret_key.txt")
        try:
            os.makedirs(cfg_dir or ".", exist_ok=True)
            if os.path.exists(key_path):
                with open(key_path, "r", encoding="utf-8") as f:
                    app.secret_key = f.read().strip() or secrets.token_hex(32)
            else:
                app.secret_key = secrets.token_hex(32)
                with open(key_path, "w", encoding="utf-8") as f:
                    f.write(app.secret_key)
        except Exception as e:
            log.error(f"secret_key bootstrap failed, using a per-restart "
                      f"key (everyone will need to re-login after every "
                      f"restart until this is fixed): {e}")
            app.secret_key = secrets.token_hex(32)

    app.permanent_session_lifetime = timedelta(hours=12)

    _PUBLIC_PATHS = {"/login", "/logout"}

    _LOGIN_HTML = """<!DOCTYPE html><html><head><meta charset="utf-8">
<title>TPNODL PSCC — {mode}</title>
<style>
  body{{margin:0;height:100vh;display:flex;align-items:center;justify-content:center;
       background:#0a0e1a;font-family:'Inter',sans-serif}}
  .card{{background:#111827;border:1px solid #1e2d4a;border-radius:10px;padding:32px 36px;
        width:300px;text-align:center}}
  h1{{color:#00c4ff;font-size:20px;margin:0 0 4px;font-family:'Rajdhani',sans-serif}}
  p.sub{{color:#8899bb;font-size:12px;margin:0 0 20px}}
  input{{width:100%;box-sizing:border-box;padding:10px 12px;border-radius:6px;
        border:1px solid #2a3f6a;background:#1a2340;color:#e8eaf6;font-size:14px;margin-bottom:12px}}
  button{{width:100%;padding:10px;border-radius:6px;border:none;background:#00c4ff;
         color:#031018;font-weight:700;font-size:14px;cursor:pointer}}
  .err{{color:#ff3d71;font-size:12px;margin:-4px 0 12px}}
  .hint{{color:#4a5a7a;font-size:11px;margin-top:14px}}
</style></head><body>
  <form class="card" method="POST">
    <h1>⚡ TPNODL PSCC</h1>
    <p class="sub">Realtime Load &amp; Voltage Monitor</p>
    {error_html}
    <input type="password" name="password" placeholder="{placeholder}" autofocus required>
    <button type="submit">{button}</button>
    {hint_html}
  </form>
</body></html>"""

    def _is_authed():
        return bool(session.get("authed"))

    @app.before_request
    def _require_login():
        path = request.path
        if path in _PUBLIC_PATHS or path.startswith("/static/"):
            return None
        pwd_hash = _cfg.get("auth.password_hash", "") if _cfg else ""
        if not pwd_hash:
            # No password set yet — first-run state. Don't lock the
            # operator out of their own freshly-installed system; send
            # them to /login, which detects this and asks them to SET one
            # instead of demanding a password that doesn't exist yet.
            if path.startswith("/api/"):
                return jsonify({"ok": False, "error": "auth_not_configured",
                                "message": "No password set yet — visit /login to set one"}), 401
            return redirect("/login")
        if _is_authed():
            return None
        if path.startswith("/api/"):
            return jsonify({"ok": False, "error": "auth_required"}), 401
        return redirect("/login")

    @app.route("/login", methods=["GET", "POST"])
    def login():
        pwd_hash = _cfg.get("auth.password_hash", "") if _cfg else ""
        first_run = not pwd_hash
        error = ""
        if request.method == "POST":
            entered = request.form.get("password", "")
            if not _AUTH_LIB_AVAILABLE:
                error = "Server misconfigured (werkzeug security module missing)"
            elif first_run:
                if len(entered) < 6:
                    error = "Choose a password with at least 6 characters"
                else:
                    _cfg.set("auth.password_hash", generate_password_hash(entered))
                    session.clear()
                    session["authed"] = True
                    session.permanent = True
                    return redirect("/")
            elif check_password_hash(pwd_hash, entered):
                session.clear()
                session["authed"] = True
                session.permanent = True
                return redirect("/")
            else:
                error = "Incorrect password"
            first_run = not _cfg.get("auth.password_hash", "")  # re-check after a failed/short attempt

        html = _LOGIN_HTML.format(
            mode="Set Password" if first_run else "Login",
            placeholder="Choose a password" if first_run else "Password",
            button="Set Password & Continue" if first_run else "Log In",
            error_html=f'<div class="err">{error}</div>' if error else "",
            hint_html='<div class="hint">First-time setup — this sets the password '
                      'for everyone accessing this dashboard.</div>' if first_run else "")
        return Response(html, mimetype="text/html")

    @app.route("/logout")
    def logout():
        session.clear()
        return redirect("/login")

    # ── System status ────────────────────────────────────
    @app.route("/api/status")
    def api_status():
        weather_anomalies = []
        if _MBSCADA_AVAILABLE:
            try:
                from modules.weather_report import get_weather_anomalies
                weather_anomalies = get_weather_anomalies()
            except Exception as e:
                log.warning(f"api_status: get_weather_anomalies failed: {e}")
        return jsonify({
            "scraper": _scraper.status(),
            "alerts":  _store.summary(),
            "feeders": len(_fm),
            "weather_anomalies": weather_anomalies,
        })

    # ── Manual fetch trigger ─────────────────────────────
    @app.route("/api/fetch", methods=["POST"])
    def api_fetch():
        def _run():
            live = _scraper.fetch()
            if live:
                _detector.run(live)
        threading.Thread(target=_run, daemon=True).start()
        return jsonify({"ok": True, "message": "Fetch started"})

    # ── Probe — inspect raw Probus API response ──────────
    @app.route("/api/probe")
    def api_probe():
        """
        Test any Probus endpoint and return its raw response.
        Usage: GET /api/probe?path=/api/login&method=POST
        """
        import requests, urllib3
        urllib3.disable_warnings()
        path   = request.args.get("path", "/api/login")
        method = request.args.get("method", "GET").upper()
        base   = "https://tpnodl.probussense.com"
        url    = base + path
        headers = {
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json",
            "Origin": base,
        }
        if _scraper._token:
            headers["Authorization"] = f"Bearer {_scraper._token}"

        try:
            if method == "POST":
                payload = {
                    "username": _cfg.get("scraper.username","1925"),
                    "password": _cfg.get("scraper.password",""),
                    "role": _cfg.get("scraper.role","Sense Admin")
                }
                r = requests.post(url, json=payload, headers=headers,
                                  verify=False, timeout=15)
            else:
                r = requests.get(url, headers=headers, verify=False, timeout=15)

            try:
                body = r.json()
            except Exception:
                body = r.text[:2000]

            return jsonify({
                "url": url,
                "method": method,
                "status": r.status_code,
                "response_headers": dict(r.headers),
                "body": body
            })
        except Exception as e:
            return jsonify({"url": url, "error": str(e)}), 500

    # ── Live data ────────────────────────────────────────
    @app.route("/api/live")
    def api_live():
        data   = _scraper.last_data
        vn     = _cfg.get("voltage.vn_kv", 33.0)
        ov_lim = vn * (1 + _cfg.get("voltage.ov_pct", 6.0) / 100)
        uv_lim = vn * (1 - _cfg.get("voltage.uv_pct", 9.0) / 100)
        ol_thr = _cfg.get("voltage.load_threshold_pct", 100.0) / 100
        off_thr= _cfg.get("voltage.feeder_off_threshold_a", 1.0)
        # Meta lookup fallback (auto from Probus)
        meta_lookup = getattr(_scraper, "_meta_lookup", {})

        enriched = []
        for r in data:
            ac = (r.get("AssetCode") or "").strip()
            vr = float(r.get("Vr") or 0)
            vy = float(r.get("Vy") or 0)
            vb = float(r.get("Vb") or 0)
            ir = float(r.get("Ir") or 0)
            iy = float(r.get("Iy") or 0)
            ib = float(r.get("Ib") or 0)
            vavg = (vr+vy+vb)/3
            imax = max(ir,iy,ib)

            # ── Org fields: Feeder Master first, meta_lookup fallback ──
            master = _fm.lookup(ac)
            if master and master.get("FeederName"):
                circle   = master.get("CircleName","")
                division = master.get("DivisionName","")
                gss      = master.get("GssName","")
                feeder   = master.get("FeederName","") or ac
                f_type   = master.get("FeederType","")
            else:
                # Fallback to meta_lookup (auto-fetched from Probus)
                meta = meta_lookup.get(ac, {})
                circle   = meta.get("Circle","")   or r.get("Circle","")
                division = meta.get("Division","") or r.get("Division","")
                gss      = meta.get("Gss","")      or r.get("Gss","")
                feeder   = meta.get("Feeder","")   or r.get("Feeder","") or ac
                f_type   = meta.get("FeederType","") or r.get("FeederType","")

            fn    = feeder.upper()
            is_bc = bool(master.get("IsBusCoupler") if master else False or
                         ("BUS" in fn and "COUPL" in fn))

            # ── Power Sign correction ─────────────────────
            # PowerSign: '' or '+' = as-is, '-' = negate (reverse CT), 'solar' = generation (keep as-is)
            raw_ap = float(r.get("ActivePower")  or 0)
            raw_sp = float(r.get("ApparentPower") or 0)
            sign = (master.get("PowerSign","") if master else "").strip().lower()
            if sign == "-":
                ap = -raw_ap   # reverse CT — negate
                sp = -raw_sp
            elif sign == "solar":
                ap = raw_ap    # solar generation — keep as-is (may be negative = exporting)
                sp = raw_sp
            else:
                ap = abs(raw_ap)   # default: all load feeders should be positive MW
                sp = abs(raw_sp)

            # ── Electrical status ─────────────────────────
            # Voltage unbalance % — used to distinguish UV from PT fault
            v_unbal = (max(abs(vr-vavg), abs(vy-vavg), abs(vb-vavg)) / vavg * 100) if vavg > 1 else 0

            # V_status: match violation.py logic exactly
            # BC excluded from UV, high unbalance = PT fault not UV
            if vavg > ov_lim:
                vs = "OV"
            elif vavg < uv_lim and vavg > 1:
                if is_bc:
                    vs = "OK"        # BC voltage unreliable for UV
                elif v_unbal > 25:
                    vs = "PT"        # PT phase missing
                elif v_unbal < 10:
                    vs = "UV"        # True balanced UV
                else:
                    vs = "OK"        # Moderate unbalance — indeterminate
            else:
                vs = "OK"
            is_  = "NC"
            load_pct = None
            if master:
                rating = float(master.get("FeederRating") or 0)
                if is_bc:
                    # BC normal state = 0A (breaker open/idle)
                    if imax < off_thr:
                        is_      = "OK"    # idle — normal state
                        load_pct = None
                    else:
                        # BC carrying load — reflect ACTUAL detection state,
                        # not just raw threshold (matches violation engine logic)
                        bc_ac = r.get("AssetCode","")
                        confirmed = bc_ac in (_detector._bc._active.values()
                                              if hasattr(_detector._bc,'_active') else {})
                        # Check if this BC has any confirmed active diversion
                        has_confirmed = any(
                            d.get("bc") == bc_ac
                            for d in getattr(_detector._bc, "_active", {}).values()
                        )
                        has_pending = any(
                            p.get("bac") == bc_ac
                            for p in getattr(_detector._bc, "_pending", {}).values()
                        )
                        if has_confirmed:
                            is_ = "DIV"        # confirmed diversion
                        elif has_pending:
                            is_ = "DIV_PEND"   # awaiting confirmation
                        else:
                            is_ = "DIV_RAW"    # current present, not yet validated
                        if rating > 10:
                            load_pct = round(imax / rating * 100, 2)
                        else:
                            load_pct = None
                elif rating > 0:
                    load_pct = round(imax / rating * 100, 2)
                    is_solar_feeder = bool((master or {}).get("IsSolarPlant"))
                    if is_solar_feeder:
                        # Solar plants use a tighter threshold + wait-time
                        # confirmation (handled by the violation engine, NOT
                        # a simple current check). 0A/low current is normal
                        # during night/cloud cover and must not show as OFF
                        # here just because it's below the standard off_thr.
                        # Defer to the actual confirmed alert state in DB —
                        # show OFF only if a FEEDER_OFF alert is genuinely
                        # active (i.e. the wait period already confirmed it).
                        ak_off_chk = f"{ac}_FEEDER_OFF"
                        if _store.is_active(ak_off_chk):
                            is_ = "OFF"
                        elif imax / rating > ol_thr:
                            is_ = "OL"
                        else:
                            is_ = "OK"
                    elif imax < off_thr:
                        is_ = "OFF"
                    elif imax / rating > ol_thr:
                        is_ = "OL"
                    else:
                        is_ = "OK"

            row = {
                **r,
                # Always override org fields with authoritative values
                "Circle":      circle,
                "Division":    division,
                "Gss":         gss,
                "Feeder":      feeder,
                "FeederType":  f_type,
                "IsBusCoupler":is_bc,
                # Computed fields
                "Vavg":        round(vavg, 4),
                "Imax":        round(imax, 4),
                "V_status":    vs,
                "I_status":    is_,
                "LoadPct":     load_pct,
                "MasterLinked":master is not None,
            }
            enriched.append(row)

        return jsonify({
            "data":       enriched,
            "fetched_at": _scraper.last_fetch_time,
            "count":      len(enriched),
            "thresholds": {"ov_limit": ov_lim, "uv_limit": uv_lim,
                           "vn": vn, "off_a": off_thr}
        })

    # ── Alerts ───────────────────────────────────────────
    @app.route("/api/alerts")
    def api_alerts():
        params = request.args
        data = _store.all(
            limit=int(params.get("limit", 300)),
            unacked_only=params.get("unacked") == "1",
            vtype=params.get("type"),
            circle=params.get("circle"),
            active_only=params.get("active") == "1",
        )
        # Refresh GSS/Feeder/Circle from current Feeder Master
        # (alerts may have been created with old names before FM was updated)
        meta_lookup = getattr(_scraper, "_meta_lookup", {})
        alert_types = _cfg.get("email.alert_types") or []  # checked = latched types
        for a in data:
            ac = a.get("AssetCode","")
            if ac and not ac.startswith("GSS_"):
                fm = _fm.lookup(ac)
                if fm and fm.get("FeederName"):
                    a["Circle"]   = fm.get("CircleName","")  or a.get("Circle","")
                    a["Division"] = fm.get("DivisionName","") or a.get("Division","")
                    a["Gss"]      = fm.get("GssName","")     or a.get("Gss","")
                    a["Feeder"]   = fm.get("FeederName","")  or a.get("Feeder","")
                else:
                    meta = meta_lookup.get(ac, {})
                    if meta:
                        a["Circle"]   = meta.get("Circle","")   or a.get("Circle","")
                        a["Division"] = meta.get("Division","") or a.get("Division","")
                        a["Gss"]      = meta.get("Gss","")      or a.get("Gss","")
                        a["Feeder"]   = meta.get("Feeder","")   or a.get("Feeder","")
            # Tag whether this type is latched (checked in UI)
            # Types that are latched (stay visible while active, show "Active Since")
            LATCHED = {"OV","UV","OL","FEEDER_OFF","LOAD_DIVERTED"}
            a["is_latched"] = (a.get("type","") in alert_types) if alert_types else (a.get("type","") in LATCHED)
        return jsonify({"data": data, "summary": _store.summary()})

    @app.route("/api/alerts/<alert_id>/force-clear", methods=["POST"])
    def api_force_clear_alert(alert_id):
        """
        Manual safety-net action: force-deactivate ANY active alert
        regardless of its type's normal auto-clear logic. Intended for
        cases where an alert is stuck active due to a code-version
        mismatch, a data gap, or any other edge case the automatic
        clearing logic hasn't covered yet — gives the operator a direct
        recovery path without waiting for a code fix to propagate.
        Body (optional): {reason: "..."}
        """
        body = request.json or {}
        reason = body.get("reason", "Manually force-cleared by operator")
        try:
            _store.clear_stale_alert(alert_id, reason=reason)
            log.warning(f"Alert {alert_id} force-cleared by operator: {reason}")
            return jsonify({"ok": True})
        except Exception as e:
            log.error(f"api_force_clear_alert failed: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/alerts/<alert_id>/ack", methods=["POST"])
    def api_ack(alert_id):
        ok = _store.ack(alert_id)

        # Special case: acknowledging a Bus Coupler FEEDER_OFF alert.
        # BC's normal idle state is 0A, so a BC FEEDER_OFF is only a real
        # concern if it's masking an ongoing diversion (BC tripped while
        # still carrying another feeder's diverted load). If every OTHER
        # feeder at the same GSS is genuinely loaded (>off_thr), the BC's
        # 0A reading is just its normal idle state — clear the alert too,
        # so it disappears from the Active list on ack instead of lingering.
        try:
            if ok:
                with _store._conn() as c:
                    row = c.execute(
                        "SELECT type, feeder, gss, is_active FROM alerts WHERE id=?",
                        (alert_id,)
                    ).fetchone()
                if row:
                    a_type, a_feeder, a_gss, a_active = row
                    feeder_name = (a_feeder or "").upper()
                    is_bc = "BUS COUPL" in feeder_name
                    if a_type == "FEEDER_OFF" and is_bc and a_active:
                        off_thr = _cfg.get("voltage.feeder_off_threshold_a", 1.0)
                        gss_rows = [
                            r for r in (_scraper.last_data or [])
                            if r.get("Gss","") == a_gss and not r.get("IsBusCoupler")
                        ]
                        if gss_rows:
                            all_loaded = all(
                                r.get("_imax", max(float(r.get("Ir",0)),
                                                    float(r.get("Iy",0)),
                                                    float(r.get("Ib",0)))) > off_thr
                                for r in gss_rows
                            )
                            if all_loaded:
                                _store.clear_stale_alert(
                                    alert_id,
                                    "Acknowledged — all feeders at GSS loaded, "
                                    "BC 0A confirmed normal idle state")
                                log.info(f"BC FEEDER_OFF {alert_id} cleared on ack "
                                         f"— all {len(gss_rows)} feeders at {a_gss} loaded")
        except Exception as e:
            log.warning(f"BC-on-ack auto-clear check failed for {alert_id}: {e}")

        return jsonify({"ok": ok})

    @app.route("/api/alerts/ack-all", methods=["POST"])
    def api_ack_all():
        _store.ack_all()
        return jsonify({"ok": True})

    @app.route("/api/bc-pending")
    def api_bc_pending():
        """Return BC diversion candidates currently under confirmation — for frontend blinking."""
        try:
            pending = _detector._bc.get_pending_state()
            return jsonify({"ok": True, "pending": pending})
        except Exception as e:
            return jsonify({"ok": False, "pending": [], "error": str(e)})

    # ── Field-Level Diversion Marking ─────────────────────────────────────
    # See modules/field_diversion.py for the full design rationale — this
    # is a manual operator annotation on top of an active FEEDER_OFF alert,
    # for diversions that happen outside the automatically-monitored Bus
    # Coupler path (e.g. a field crew physically jumpering load to a
    # different feeder, possibly at a different GSS).
    @app.route("/api/field-diversion/options")
    def api_fd_options():
        """Dropdown options for the 'select source GSS & Feeder' picker."""
        if not _FIELD_DIVERSION_AVAILABLE:
            return jsonify({"ok": False, "error": "field_diversion module not available"}), 503
        try:
            opts = _fd_picker_options(_fm)
            return jsonify({"ok": True, "options": opts})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/field-diversion/mark", methods=["POST"])
    def api_fd_mark():
        """
        Mark an active FEEDER_OFF alert as field-diverted.
        Body: {alert_id, asset_code, feeder_name, gss_name,
               source_asset_code, source_feeder_name, source_gss_name,
               diversion_time (optional, ISO string — when the diversion
               actually happened; defaults to 'now' if omitted),
               note (optional), marked_by (optional)}
        """
        if not _FIELD_DIVERSION_AVAILABLE:
            return jsonify({"ok": False, "error": "field_diversion module not available"}), 503
        body = request.json or {}
        required = ("alert_id", "asset_code", "source_asset_code")
        missing = [k for k in required if not body.get(k)]
        if missing:
            return jsonify({"ok": False, "error": f"Missing required field(s): {missing}"}), 400
        try:
            # Validate the alert is actually active before allowing the mark —
            # prevents marking an alert that already cleared (stale UI state)
            if not _store.is_active(
                f"{body['asset_code']}_FEEDER_OFF"
            ):
                return jsonify({"ok": False,
                                "error": "This alert is no longer active — cannot mark"}), 409

            result = _fd_mark(
                alert_id           = body["alert_id"],
                asset_code         = body["asset_code"],
                feeder_name        = body.get("feeder_name",""),
                gss_name           = body.get("gss_name",""),
                source_asset_code  = body["source_asset_code"],
                source_feeder_name = body.get("source_feeder_name",""),
                source_gss_name    = body.get("source_gss_name",""),
                diversion_time     = body.get("diversion_time",""),
                note               = body.get("note",""),
                marked_by          = body.get("marked_by",""),
            )
            return jsonify({"ok": True, "mark": result})
        except Exception as e:
            log.error(f"api_fd_mark failed: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/field-diversion/clear", methods=["POST"])
    def api_fd_clear():
        """Un-mark a previously field-diversion-marked alert.
        Body: {alert_id, reason (optional)}"""
        if not _FIELD_DIVERSION_AVAILABLE:
            return jsonify({"ok": False, "error": "field_diversion module not available"}), 503
        body = request.json or {}
        alert_id = body.get("alert_id","")
        if not alert_id:
            return jsonify({"ok": False, "error": "alert_id is required"}), 400
        try:
            cleared = _fd_clear(alert_id, reason=body.get("reason","manual unmark"))
            return jsonify({"ok": True, "cleared": cleared})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/field-diversion/active")
    def api_fd_active():
        """All currently active field-diversion marks — used by the
        dashboard/Alert Log to annotate matching alerts, and by
        mgmt_report.py to include them in the management report."""
        if not _FIELD_DIVERSION_AVAILABLE:
            return jsonify({"ok": True, "marks": []})  # graceful, not an error
        try:
            marks = _fd_get_all()
            return jsonify({"ok": True, "marks": marks})
        except Exception as e:
            return jsonify({"ok": False, "marks": [], "error": str(e)})

    @app.route("/api/notify_queue/status")
    def api_notify_queue_status():
        """Diagnostic view into the persistent notification retry queue
        (see notify_queue.py) — shows pending/sent/failed counts plus the
        most recent permanent failures, so a stuck/exhausted delivery is
        visible on the dashboard instead of only in logs."""
        if not _notify_queue:
            return jsonify({"ok": True, "enabled": False,
                            "message": "notify_queue not wired into app.py — "
                                       "sends go direct, no retry on failure"})
        return jsonify({
            "ok": True, "enabled": True,
            "summary": _notify_queue.status_summary(),
            "recent_failures": _notify_queue.recent_failures(),
        })

    # ── MB SCADA Weather Login Config ──────────────────────────────────────
    @app.route("/api/weather/anomalies")
    def api_weather_anomalies():
        """
        On-screen-notification feed for suspected weather-sensor faults
        (exact-zero readings that DID report, or a frozen/non-communicating
        sensor). Intentionally NOT shown in the management report — see
        weather_report.build_weather_html_table()'s docstring. The
        dashboard should poll this (or just read the same field already
        included in /api/status) and surface it as a toast/banner.
        """
        if not _MBSCADA_AVAILABLE:
            return jsonify({"ok": True, "anomalies": []})
        try:
            from modules.weather_report import get_weather_anomalies
            return jsonify({"ok": True, "anomalies": get_weather_anomalies()})
        except Exception as e:
            log.error(f"api_weather_anomalies failed: {e}")
            return jsonify({"ok": False, "anomalies": [], "error": str(e)})

    @app.route("/api/weather/config", methods=["GET"])
    def api_weather_config_get():
        """
        Return saved weather credentials — PASSWORD NEVER ECHOED BACK
        (write-only field, same pattern as the main config's password
        masking). Only username/company populate the Settings UI on load.
        """
        try:
            path = os.path.join("data", "weather_config.json")
            if not os.path.exists(path):
                return jsonify({"ok": True, "mbscada_user": "", "mbscada_company": "TPNODL"})
            with open(path) as f:
                cfg = json.load(f)
            return jsonify({
                "ok": True,
                "mbscada_user":    cfg.get("mbscada_user",""),
                "mbscada_company": cfg.get("mbscada_company","TPNODL"),
                # mbscada_password intentionally omitted
            })
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/weather/config", methods=["POST"])
    def api_weather_config_post():
        """
        Save MB SCADA weather login credentials to data/weather_config.json
        — the exact file/keys mbscada_scraper.py's _load_credentials()
        already expects, so no changes needed to that module.
        Body: {mbscada_user, mbscada_password, mbscada_company}
        """
        body = request.json or {}
        username = body.get("mbscada_user","").strip()
        company  = body.get("mbscada_company","TPNODL").strip() or "TPNODL"
        password = body.get("mbscada_password","")

        if not username:
            return jsonify({"ok": False, "error": "Username is required"}), 400

        try:
            os.makedirs("data", exist_ok=True)
            path = os.path.join("data", "weather_config.json")
            # Preserve existing password if a blank one was submitted —
            # mirrors the main config's "leave blank to keep existing"
            # password-field convention, so re-saving username/company
            # alone doesn't accidentally wipe a previously-saved password.
            existing_password = ""
            if os.path.exists(path):
                try:
                    with open(path) as f:
                        existing_password = json.load(f).get("mbscada_password","")
                except Exception:
                    pass
            final_password = password if password else existing_password

            with open(path, "w") as f:
                json.dump({
                    "mbscada_user": username,
                    "mbscada_password": final_password,
                    "mbscada_company": company,
                }, f, indent=2)

            log.info(f"Weather config saved (user={username}, company={company})")

            # Kick the background loop to pick up new credentials on its
            # next cycle — invalidate any cached JWT so it re-authenticates
            # immediately with the new credentials rather than waiting up
            # to 85 minutes for the old cached token to expire.
            if _MBSCADA_AVAILABLE:
                try:
                    with _mb._jwt['lock']:
                        _mb._jwt['token'] = None
                except Exception:
                    pass

            return jsonify({"ok": True})
        except Exception as e:
            log.error(f"api_weather_config_post failed: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/weather/test", methods=["POST"])
    def api_weather_test():
        """
        Test connectivity for one site (or the first configured site).
        The block-data API confirmed to need NO authentication in testing —
        this just fetches today's data for a real siteId and reports back
        whether data came back, matching what the Settings UI's per-site
        'Test' button needs.
        Body (optional): {site_id: "1004_01"}
        """
        if not _MBSCADA_AVAILABLE:
            return jsonify({"ok": False, "error": "mbscada_scraper module not installed on this server"}), 503
        body = request.json or {}
        site_id = body.get("site_id")
        try:
            result = _mb.test_connection(site_id)
            status_code = 200 if result.get("ok") else 502
            return jsonify(result), status_code
        except Exception as e:
            log.error(f"api_weather_test failed: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    # ── Weather Station Site Management ─────────────────────────────────────
    # Matches the STLF module's "WMS Weather Station Configuration" UI
    # pattern — sites are fully user-managed, no code change required to
    # add/remove a physical weather station.
    @app.route("/api/weather/sites", methods=["GET"])
    def api_weather_sites_get():
        """List all configured weather stations."""
        if not _MBSCADA_AVAILABLE:
            return jsonify({"ok": True, "sites": []})
        try:
            sites = _mb.load_sites()
            return jsonify({"ok": True, "sites": sites})
        except Exception as e:
            return jsonify({"ok": False, "sites": [], "error": str(e)})

    @app.route("/api/weather/sites", methods=["POST"])
    def api_weather_sites_save():
        """
        Save the FULL site list (replaces entirety — same pattern as
        Feeder Master's bulk save). Body: {sites: [...]}
        Each site: {site_id, sheet_name, display_name, location, circle,
                    weight, active}
        """
        if not _MBSCADA_AVAILABLE:
            return jsonify({"ok": False, "error": "mbscada_scraper module not available"}), 503
        body = request.json or {}
        sites = body.get("sites", [])
        # Basic validation — every site needs at minimum a site_id
        for s in sites:
            if not s.get("site_id"):
                return jsonify({"ok": False, "error": "Every site must have a site_id"}), 400
        try:
            ok = _mb.save_sites(sites)
            if ok:
                return jsonify({"ok": True, "count": len(sites)})
            return jsonify({"ok": False, "error": "Failed to save sites file"}), 500
        except Exception as e:
            log.error(f"api_weather_sites_save failed: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/weather/sites/test", methods=["POST"])
    def api_weather_site_test():
        """Per-site 'Test' button — body: {site_id}"""
        if not _MBSCADA_AVAILABLE:
            return jsonify({"ok": False, "error": "mbscada_scraper module not available"}), 503
        body = request.json or {}
        site_id = body.get("site_id")
        if not site_id:
            return jsonify({"ok": False, "error": "site_id is required"}), 400
        try:
            result = _mb.test_connection(site_id)
            status_code = 200 if result.get("ok") else 502
            return jsonify(result), status_code
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/weather/live")
    def api_weather_live():
        """Current circle-wise weather (cached, refreshed by the background
        loop) — used by the Settings UI's coverage map and any dashboard
        widget that wants live weather without waiting for a report."""
        if not _MBSCADA_AVAILABLE:
            return jsonify({"ok": True, "circle_weather": {}})
        try:
            cached = _mb.get_cache()
            cw = cached.get("circle_weather") or _mb.get_circle_weather_now()
            return jsonify({"ok": True, "circle_weather": cw})
        except Exception as e:
            return jsonify({"ok": False, "circle_weather": {}, "error": str(e)})

    @app.route("/api/alerts/clear", methods=["POST"])
    def api_clear_alerts():
        _store.clear()
        return jsonify({"ok": True})

    @app.route("/api/alerts/force-clear-type", methods=["POST"])
    def api_force_clear_type():
        """
        Force-deactivate ALL active alerts of a given type.
        POST body: {"type": "LOAD_DIVERTED", "reason": "bug cleanup"}
        Used for post-bug manual cleanup when auto-clear can't fire.
        """
        data   = request.get_json(silent=True) or {}
        vtype  = data.get("type", "").strip()
        reason = data.get("reason", "manual API call")
        if not vtype:
            return jsonify({"ok": False, "error": "type is required"}), 400
        n = _store.force_clear_by_type(vtype, reason)
        return jsonify({"ok": True, "cleared": n, "type": vtype})

    @app.route("/api/alerts/<alert_id>/force-clear", methods=["POST"])
    def api_force_clear_one(alert_id):
        """Force-deactivate a single alert by ID regardless of its current condition."""
        data   = request.get_json(silent=True) or {}
        reason = data.get("reason", "manual API call")
        ok = _store.force_clear_alert(alert_id, reason)
        return jsonify({"ok": ok, "id": alert_id})

    @app.route("/api/alerts/fix-instant-types", methods=["POST"])
    def api_fix_instant_types():
        """One-time fix: mark LOAD_RESTORED / SLD / SLR alerts as is_active=0."""
        import sqlite3 as _sq
        INSTANT = ("LOAD_RESTORED","SUDDEN_LOAD_DROP","SUDDEN_LOAD_RAISE")
        with _sq.connect("data/alerts.db") as c:
            n = c.execute(
                "UPDATE alerts SET is_active=0 WHERE type IN ({}) AND is_active=1".format(
                    ",".join("?"*len(INSTANT))), INSTANT).rowcount
        return jsonify({"ok": True, "fixed": n})

    @app.route("/api/alerts/cleanup-stale", methods=["POST"])
    def api_alerts_cleanup_stale():
        live_data   = _scraper.last_data
        live_assets = {r.get("AssetCode","") for r in live_data}
        off_thr     = _cfg.get("voltage.feeder_off_threshold_a", 1.0)
        meta        = getattr(_scraper, "_meta_lookup", {})
        deactivated = []

        # Build live imax lookup
        live_imax = {}
        for r in live_data:
            ac = r.get("AssetCode","")
            if ac:
                live_imax[ac] = max(float(r.get("Ir",0)),
                                    float(r.get("Iy",0)),
                                    float(r.get("Ib",0)))

        # Build current GSS set
        current_gss = set()
        for entry in _fm.all():
            if entry.get("GssName"):
                current_gss.add(entry["GssName"].strip())
        for m in meta.values():
            if m.get("Gss"):
                current_gss.add(m["Gss"].strip())

        for a in _store.all(active_only=True, limit=500):
            ac    = a.get("AssetCode","")
            atype = a.get("type","")
            gss   = (a.get("Gss") or "").strip()

            if ac.startswith("GSS_"):
                if gss and gss not in current_gss:
                    _store.clear_stale_alert(a["id"], "GSS renamed/removed")
                    deactivated.append({"id":a["id"], "gss":gss, "reason":"GSS renamed"})

            elif atype == "FEEDER_OFF":
                curr = live_imax.get(ac)
                if curr is not None and curr > off_thr:
                    _store.clear_stale_alert(a["id"], f"Feeder running at {curr:.2f}A")
                    deactivated.append({"id":a["id"], "asset":ac, "reason":f"Feeder ON {curr:.2f}A"})

            elif atype == "LOAD_DIVERTED":
                # Do NOT clear here — clearing LOAD_DIVERTED requires checking
                # BOTH feeder restoration AND BC load reduction (accounting for
                # any other feeders sharing the same BC). That dual-check logic
                # lives in violation.py's BusCouplerDiversionDetector and runs
                # automatically every fetch cycle. Manual "Cleanup Stale" must
                # not bypass it, or a diversion could be cleared while BC is
                # still genuinely carrying the diverted load.
                pass

            elif atype in ("OV","UV","OL"):
                if ac and ac not in live_assets:
                    _store.clear_stale_alert(a["id"], "Meter not in live data")
                    deactivated.append({"id":a["id"], "asset":ac, "reason":"Meter offline"})

        log.info(f"Stale cleanup: {len(deactivated)} deactivated")
        return jsonify({"ok":True, "deactivated":len(deactivated), "details":deactivated})


    @app.route("/api/alerts/export")
    def api_alerts_export():
        """
        Export alerts as CSV or JSON.
        Params: start=YYYY-MM-DD, end=YYYY-MM-DD,
                types=OV,UV,OL (comma-sep, optional),
                fmt=csv|json (default csv)
        """
        import csv, io
        start  = request.args.get("start", "2026-01-01")
        end    = request.args.get("end",   datetime.now().strftime("%Y-%m-%d"))
        types  = request.args.get("types","")
        fmt    = request.args.get("fmt","csv")
        vtypes = [t.strip() for t in types.split(",") if t.strip()] or None

        data = _store.export(start, end, vtypes)

        if fmt == "json":
            from flask import Response
            return Response(
                json.dumps({"count": len(data), "data": data}, indent=2, default=str),
                mimetype="application/json",
                headers={"Content-Disposition":
                         f"attachment;filename=tpnodl_alerts_{start}_{end}.json"}
            )
        # CSV
        if not data:
            csv_str = "No data found for selected range\n"
        else:
            COLS = ["type","severity","AssetCode","Feeder","Circle","Division","Gss",
                    "FeederType","value","limit","detail",
                    "Vr","Vy","Vb","Ir","Iy","Ib","FeederRating",
                    "first_seen","last_seen","cleared_at","duration_s",
                    "acked","notified_email","notified_wa"]
            buf = io.StringIO()
            writer = csv.DictWriter(buf, fieldnames=COLS, extrasaction="ignore")
            writer.writeheader()
            for row in data:
                # Format duration as plain text to prevent Excel time misinterpretation
                # e.g. "1h 23m 45s" instead of "01:23:45" which Excel reads as time
                dur = row.get("duration_s")
                if dur is not None:
                    try:
                        dur = float(dur)
                        m, s = divmod(int(dur), 60)
                        h, m = divmod(m, 60)
                        if h:
                            row["duration_s"] = f"{h}h {m:02d}m {s:02d}s"
                        elif m:
                            row["duration_s"] = f"{m}m {s:02d}s"
                        else:
                            row["duration_s"] = f"{s}s"
                    except Exception:
                        pass  # leave as-is if not a number
                writer.writerow(row)
            csv_str = buf.getvalue()

        from flask import Response
        return Response(
            csv_str,
            mimetype="text/csv",
            headers={"Content-Disposition":
                     f"attachment;filename=tpnodl_alerts_{start}_{end}.csv"}
        )

    # ── Feeder Master CRUD ───────────────────────────────
    @app.route("/api/feeders")
    def api_feeders():
        circle = request.args.get("circle")
        data = _fm.all()
        if circle:
            data = [f for f in data if f.get("CircleName","").upper() == circle.upper()]
        return jsonify({"data": data, "count": len(data)})

    @app.route("/api/feeders", methods=["POST"])
    def api_feeder_add():
        body = request.json or {}
        if not body.get("FeederName"):
            return jsonify({"ok": False, "error": "FeederName required"}), 400
        entry = _fm.add(body)
        return jsonify({"ok": True, "entry": entry})

    @app.route("/api/feeders/<int:idx>", methods=["PUT"])
    def api_feeder_update(idx):
        body = request.json or {}
        entry = _fm.update(idx, body)
        if entry is None:
            return jsonify({"ok": False, "error": "Index out of range"}), 404
        return jsonify({"ok": True, "entry": entry})

    @app.route("/api/feeders/<int:idx>", methods=["DELETE"])
    def api_feeder_delete(idx):
        ok = _fm.delete(idx)
        return jsonify({"ok": ok})

    @app.route("/api/feeders/import", methods=["POST"])
    def api_feeder_import():
        body = request.json or {}
        entries = body.get("entries", [])
        result = _fm.import_bulk(entries)
        return jsonify({"ok": True, **result})

    @app.route("/api/feeders/upload-excel", methods=["POST"])
    def api_feeder_upload_excel():
        """
        Upload an Excel file to import feeder master.
        Accepts .xlsx with columns (any order, case-insensitive):
          AssetCode, Circle/CircleName, Division/DivisionName,
          GSS/GssName, Feeder/FeederName, Rating/FeederRating,
          VoltageRating, FeederType, FeederCode (all optional except AssetCode or FeederName)
        """
        if "file" not in request.files:
            return jsonify({"ok": False, "error": "No file uploaded"}), 400

        f = request.files["file"]
        if not f.filename:
            return jsonify({"ok": False, "error": "Empty filename"}), 400

        ext = f.filename.rsplit(".",1)[-1].lower()
        if ext not in ("xlsx","xls","csv"):
            return jsonify({"ok": False, "error": "Only .xlsx, .xls or .csv accepted"}), 400

        try:
            import io
            data = f.read()

            if ext == "csv":
                import csv as _csv
                reader = _csv.DictReader(io.StringIO(data.decode("utf-8-sig")))
                rows = list(reader)
            else:
                import openpyxl
                wb   = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
                ws   = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [v for v in row])))

            # Column name normalisation map
            COL_MAP = {
                "assetcode":"AssetCode","asset code":"AssetCode","meterid":"AssetCode",
                "meter id":"AssetCode","meter no":"AssetCode",
                "circle":"CircleName","circlename":"CircleName",
                "division":"DivisionName","divisionname":"DivisionName",
                "gss":"GssName","gssname":"GssName","substation":"GssName",
                "feeder":"FeederName","feedername":"FeederName","feeder name":"FeederName",
                "rating":"FeederRating","feederrating":"FeederRating",
                "rating(a)":"FeederRating","feederrating(a)":"FeederRating",
                "voltagekv":"VoltageRating","voltagerating":"VoltageRating",
                "feedertype":"FeederType","type":"FeederType",
                "feedercode":"FeederCode","code":"FeederCode",
            }

            def norm_key(k):
                return COL_MAP.get(str(k or "").lower().replace(" ","").replace("_",""), k)

            entries = []
            skipped = 0
            for row in rows:
                # Normalise column names
                nr = {norm_key(k): v for k, v in row.items() if v is not None}
                feeder_name = str(nr.get("FeederName") or "").strip()
                if not feeder_name:
                    skipped += 1
                    continue
                ac = str(nr.get("AssetCode") or "").strip()
                try:
                    rating = float(str(nr.get("FeederRating") or 200).replace("A","").strip())
                except Exception:
                    rating = 200
                try:
                    vn = float(str(nr.get("VoltageRating") or 33).replace("kV","").strip())
                except Exception:
                    vn = 33
                fn = feeder_name.upper()
                entry = {
                    "AssetCode":    ac,
                    "CircleName":   str(nr.get("CircleName")   or "").strip(),
                    "DivisionName": str(nr.get("DivisionName") or "").strip(),
                    "GssName":      str(nr.get("GssName")      or "").strip(),
                    "FeederName":   feeder_name,
                    "FeederRating": rating,
                    "VoltageRating":vn,
                    "FeederType":   str(nr.get("FeederType")   or "Non-Priority").strip(),
                    "FeederCode":   str(nr.get("FeederCode")   or "").strip(),
                    "IsBusCoupler": "BUS" in fn and "COUPL" in fn,
                }
                entries.append(entry)

            result = _fm.import_bulk(entries)
            # Rebuild meta_lookup from feeder master so dashboard picks it up
            _scraper.invalidate_meta()
            log.info(f"Excel feeder import: {result['added']} added, {result['updated']} updated, {skipped} skipped")
            return jsonify({
                "ok":      True,
                "added":   result["added"],
                "updated": result["updated"],
                "skipped": skipped,
                "total":   len(entries),
                "message": f"Imported {len(entries)} feeders ({result['added']} new, {result['updated']} updated)"
            })

        except Exception as e:
            log.error(f"Excel feeder upload error: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/feeders/template")
    def api_feeder_template():
        """Download a blank Excel template for feeder master upload."""
        import io, openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import Response

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FeederMaster"

        headers = ["AssetCode","CircleName","DivisionName","GssName",
                   "FeederName","FeederRating(A)","VoltageRating(kV)","FeederType","FeederCode"]
        # Style header
        hdr_fill = PatternFill("solid", fgColor="1a2340")
        hdr_font = Font(color="00c4ff", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(len(h)+4, 18)

        # Sample rows from current feeder master
        current = _fm.all()[:5]
        for ridx, fm_row in enumerate(current, 2):
            sample = [fm_row.get("AssetCode",""), fm_row.get("CircleName",""),
                      fm_row.get("DivisionName",""), fm_row.get("GssName",""),
                      fm_row.get("FeederName",""), fm_row.get("FeederRating",200),
                      fm_row.get("VoltageRating",33), fm_row.get("FeederType","Non-Priority"),
                      fm_row.get("FeederCode","")]
            for col, v in enumerate(sample, 1):
                ws.cell(row=ridx, column=col, value=v)

        # Add a help note
        ws.cell(row=len(current)+3, column=1,
                value="Note: AssetCode is the Probus meter serial (e.g. TPN61655). FeederRating in Amps.")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=FeederMaster_Template.xlsx"})

    # ── Config ───────────────────────────────────────────
    @app.route("/api/config")
    def api_config_get():
        return jsonify(_cfg.all())  # passwords masked

    @app.route("/api/config/<section>", methods=["POST"])
    def api_config_set(section):
        body = request.json or {}
        valid = ("scraper","voltage","email","whatsapp","contacts","notify_delay")
        if section not in valid:
            return jsonify({"ok": False, "error": f"Invalid section. Use: {valid}"}), 400
        if section == "notify_delay":
            # Flat keys — set each individually
            for k, v in body.items():
                _cfg.set(k, v)
        else:
            _cfg.update_section(section, body)
        return jsonify({"ok": True, "section": section})

    # ── Email ────────────────────────────────────────────
    @app.route("/api/email/test", methods=["POST"])
    def api_email_test():
        result = _email.send_test()
        return jsonify(result)

    @app.route("/api/email/ping", methods=["POST"])
    def api_email_ping():
        """Test TCP reachability of SMTP server (without sending email)."""
        import socket
        host = _cfg.get("email.smtp_host", "smtp.gmail.com")
        port = int(_cfg.get("email.smtp_port", 587))
        try:
            s = socket.create_connection((host, port), timeout=8)
            s.close()
            return jsonify({"ok": True,
                            "message": f"✅ {host}:{port} reachable"})
        except Exception as e:
            return jsonify({"ok": False,
                            "message": f"❌ Cannot reach {host}:{port} — {e}\n"
                                       f"Try port 465 (SSL) or check firewall/proxy"})

    # ── WhatsApp ─────────────────────────────────────────
    @app.route("/api/whatsapp/test", methods=["POST"])
    def api_wa_test():
        result = _wa.send_test()
        return jsonify(result)

    @app.route("/api/whatsapp/diagnose")
    def api_wa_diagnose():
        """Diagnostic: check chromedriver paths, ports, driver state."""
        import glob, shutil, socket as _sock
        home = os.path.expanduser("~")
        cfg_path = _cfg.get("whatsapp.chromedriver_path", "")
        ports = {}
        for p in (9222, 9223, 9224):
            try:
                s = _sock.create_connection(("127.0.0.1", p), timeout=1)
                s.close()
                ports[p] = "OPEN"
            except Exception:
                ports[p] = "closed"
        paths = []
        if cfg_path:
            paths.append({"path": cfg_path, "exists": os.path.exists(cfg_path), "source": "config"})
        sys_cd = shutil.which("chromedriver")
        if sys_cd:
            paths.append({"path": sys_cd, "exists": True, "source": "PATH"})
        for p in sorted(glob.glob(os.path.join(home, ".cache", "selenium", "chromedriver", "win64", "*", "chromedriver.exe")), reverse=True):
            paths.append({"path": p, "exists": os.path.exists(p), "source": "selenium-cache"})
        for p in sorted(glob.glob(os.path.join(home, ".wdm", "drivers", "chromedriver", "win64", "*", "chromedriver-win64", "chromedriver.exe")), reverse=True):
            paths.append({"path": p, "exists": os.path.exists(p), "source": "wdm-cache"})
        return jsonify({
            "driver_state": {"initialized": _wa._driver is not None, "ready": _wa._ready},
            "debug_ports": ports,
            "chromedriver_paths": paths,
            "wa_xpath_file": os.path.exists("data/wa_xpaths.json"),
            "wa_profile_dir": os.path.exists("data/wa_chrome_profile"),
        })

    @app.route("/api/whatsapp/test-launch", methods=["POST"])
    def api_wa_test_launch():
        """Synchronous Chrome launch test — returns exact error immediately."""
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.chrome.service import Service

            cfg_path = _cfg.get("whatsapp.chromedriver_path", "")
            if not cfg_path or not os.path.exists(cfg_path):
                return jsonify({"ok": False, "error": f"chromedriver not found at: {cfg_path}"})

            opts = Options()
            opts.add_argument("--no-sandbox")
            opts.add_argument("--headless=new")   # headless just to test driver works
            opts.add_argument("--disable-gpu")

            svc = Service(executable_path=cfg_path)
            driver = webdriver.Chrome(service=svc, options=opts)
            ver = driver.capabilities.get("browserVersion","?")
            cd_ver = driver.capabilities.get("chrome",{}).get("chromedriverVersion","?")
            driver.quit()
            return jsonify({"ok": True,
                            "chrome_version": ver,
                            "chromedriver_version": cd_ver,
                            "message": f"✅ Chrome {ver} launched OK"})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)})

    @app.route("/api/whatsapp/detect-xpaths", methods=["POST"])
    def api_wa_detect_xpaths():
        """Re-detect WhatsApp Web DOM XPaths automatically."""
        try:
            result = _wa.detect_xpaths()
            return jsonify(result)
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)})

    @app.route("/api/whatsapp/auto-connect", methods=["POST"])
    def api_wa_auto_connect():
        result = _wa.auto_connect()
        log.info(f"WA auto-connect: {'OK' if result.get('ok') else 'FAILED'}")
        return jsonify(result)

    @app.route("/api/whatsapp/status")
    def api_wa_status():
        return jsonify(_wa.get_status())

    @app.route("/api/whatsapp/init", methods=["POST"])
    def api_wa_init():
        result = _wa.open_browser()
        return jsonify(result)

    @app.route("/api/whatsapp/close", methods=["POST"])
    def api_wa_close():
        _wa.close()
        return jsonify({"ok": True})

    @app.route("/api/whatsapp/link")
    def api_wa_link():
        number = request.args.get("number","")
        message = request.args.get("message","TPNODL Monitor test")
        link = _wa.get_walink(number, message)
        return jsonify({"link": link})

    # ── Management Hourly Report ──────────────────────────
    @app.route("/api/mgmt-report/config", methods=["GET"])
    def api_mgmt_cfg_get():
        return jsonify({
            "enabled":          _cfg.get("mgmt_report.enabled", False),
            "interval_min":     _cfg.get("mgmt_report.interval_min", 60),
            "email_recipients": _cfg.get("mgmt_report.email_recipients", []),
            "wa_recipients":    _cfg.get("mgmt_report.wa_recipients", []),
            "from_addr":        _cfg.get("mgmt_report.from_addr", ""),
            "smtp_host":        _cfg.get("mgmt_report.smtp_host", ""),
            "smtp_port":        _cfg.get("mgmt_report.smtp_port", 587),
            "tls":              _cfg.get("mgmt_report.tls", "STARTTLS"),
            "password":         "***" if _cfg.get("mgmt_report.password") else "",
            "feeder_whitelist": _cfg.get("mgmt_report.feeder_whitelist", []),
            "feeder_blacklist":  _cfg.get("mgmt_report.feeder_blacklist", []),
        })

    @app.route("/api/mgmt-report/config", methods=["POST"])
    def api_mgmt_cfg_post():
        data = request.get_json() or {}
        if data.get("password") == "***":
            data.pop("password")
        for k, v in data.items():
            _cfg.set(f"mgmt_report.{k}", v)
        return jsonify({"ok": True})

    @app.route("/api/mgmt-report/send-now", methods=["POST"])
    def api_mgmt_send_now():
        if not _mgmt:
            return jsonify({"ok": False, "error": "Reporter not initialized"})
        ok = _mgmt.send_now()
        return jsonify({"ok": ok, "message": "Report sent" if ok else "Send failed — check log"})

    # ── Logs ─────────────────────────────────────────────
    @app.route("/api/logs")
    def api_logs():
        log_file = "logs/monitor.log"
        lines = []
        if os.path.exists(log_file):
            try:
                with open(log_file,"r",encoding="utf-8") as f:
                    lines = f.readlines()[-200:]
            except Exception:
                pass
        return jsonify({"lines": [l.rstrip() for l in lines]})

    log.info("All API routes registered")

    # ── Circle-wise demand summary ────────────────────────────
    @app.route("/api/demand-summary")
    def api_demand_summary():
        import os, json as _json
        from datetime import datetime, timedelta

        data = _scraper.last_data
        if not data:
            return jsonify({"circles": [], "fetched_at": None})

        circles = {}
        for r in data:
            circ = r.get("Circle","").strip()
            if not circ or circ == "—":
                continue
            raw_ap = float(r.get("ActivePower")   or 0)
            raw_sp = float(r.get("ApparentPower") or 0)
            _fm_r  = _fm.lookup((r.get("AssetCode") or "").strip())
            _sign  = (_fm_r.get("PowerSign","") if _fm_r else "").strip().lower()
            if _sign == "-":
                ap, sp = -raw_ap, -raw_sp
            elif _sign == "solar":
                ap, sp = raw_ap, raw_sp
            else:
                ap, sp = abs(raw_ap), abs(raw_sp)
            if circ not in circles:
                circles[circ] = {"Circle": circ, "MW_now": 0, "MVA_now": 0,
                                 "MW_prev": None, "MVA_prev": None,
                                 "feeder_count": 0, "ol_count": 0}
            circles[circ]["MW_now"]  += ap
            circles[circ]["MVA_now"] += sp
            circles[circ]["feeder_count"] += 1

        alerts = _store.all(limit=500, active_only=True)
        for a in alerts:
            if a.get("type") == "OL":
                circ = a.get("Circle","")
                if circ in circles:
                    circles[circ]["ol_count"] += 1

        for c in circles.values():
            c["MW_now"]  = round(c["MW_now"],  3)
            c["MVA_now"] = round(c["MVA_now"], 3)

        # Previous day same 15-min slot — from SQLite circle_15min table
        if _peak:
            for circ, cv in circles.items():
                prev = _peak.get_circle_prevday_slot(circ)
                if prev:
                    cv["MW_prev"]       = prev["mw"]
                    cv["MVA_prev"]      = prev["mva"]
                    cv["prev_slot"]     = prev["time_slot"]
                    cv["prev_label"]    = "Yesterday " + prev["time_slot"]
                else:
                    # Fallback: try ±15 min slots
                    for offset in (15, -15, 30, -30):
                        prev = _peak.get_circle_prevday_slot(circ, offset)
                        if prev:
                            cv["MW_prev"]    = prev["mw"]
                            cv["MVA_prev"]   = prev["mva"]
                            cv["prev_slot"]  = prev["time_slot"]
                            cv["prev_label"] = f"Yesterday {prev['time_slot']}"
                            break

        # Peak loads from DB
        peaks = _peak.get_current() if _peak else {}
        daily_peaks   = peaks.get("daily",   {})
        monthly_peaks = peaks.get("monthly", {})
        for circ, cv in circles.items():
            cv["peak_daily_mw"]    = round(daily_peaks.get(circ,  {}).get("peak_mw",  cv["MW_now"]),3)
            cv["peak_daily_mva"]   = round(daily_peaks.get(circ,  {}).get("peak_mva", cv["MVA_now"]),3)
            cv["peak_daily_time"]  = daily_peaks.get(circ, {}).get("peak_time","")
            cv["peak_month_mw"]    = round(monthly_peaks.get(circ,{}).get("peak_mw",  cv["MW_now"]),3)
            cv["peak_month_mva"]   = round(monthly_peaks.get(circ,{}).get("peak_mva", cv["MVA_now"]),3)
            cv["peak_month_time"]  = monthly_peaks.get(circ, {}).get("peak_time","")

        total_mw  = round(sum(c["MW_now"]  for c in circles.values()), 3)
        total_mva = round(sum(c["MVA_now"] for c in circles.values()), 3)

        return jsonify({
            "circles":               sorted(circles.values(), key=lambda x: -x["MW_now"]),
            "fetched_at":            _scraper.last_fetch_time,
            "total_mw":              total_mw,
            "total_mva":             total_mva,
            "total_peak_daily_mw":   peaks.get("total_daily_mw",  total_mw),
            "total_peak_daily_mva":  peaks.get("total_daily_mva", total_mva),
            "total_peak_daily_time": peaks.get("total_daily_peak_time", ""),
            "total_peak_month_mw":   peaks.get("total_monthly_mw",  total_mw),
            "total_peak_month_mva":  peaks.get("total_monthly_mva", total_mva),
            "total_peak_month_time": peaks.get("total_monthly_peak_time", ""),
            "peak_date":             peaks.get("date",""),
            "peak_month":            peaks.get("month",""),
        })

    @app.route("/api/peak/export")
    def api_peak_export():
        """Export monthly peaks. ?start=YYYY-MM&end=YYYY-MM&fmt=csv|json"""
        import csv, io
        now   = datetime.now()
        start = request.args.get("start", now.strftime("%Y-%m"))
        end   = request.args.get("end",   now.strftime("%Y-%m"))
        fmt   = request.args.get("fmt",  "csv")
        if not _peak:
            return jsonify({"error": "Peak store not initialized"}), 500
        data = _peak.export(start, end)
        if fmt == "json":
            from flask import Response
            return Response(json.dumps(data, indent=2, default=str),
                mimetype="application/json",
                headers={"Content-Disposition": f"attachment;filename=tpnodl_peaks_{start}_{end}.json"})
        buf = io.StringIO()
        wr  = csv.DictWriter(buf, fieldnames=["month","circle","peak_mw","peak_mva","peak_date","peak_time"])
        wr.writeheader()
        wr.writerows(data.get("monthly_by_circle",[]))
        buf.write("\nTOTAL\n")
        tw = csv.DictWriter(buf, fieldnames=["period","peak_mw","peak_mva","peak_time"])
        tw.writeheader()
        tw.writerows(data.get("monthly_total",[]))
        from flask import Response
        return Response(buf.getvalue(), mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename=tpnodl_peaks_{start}_{end}.csv"})

    @app.route("/api/peak/reset", methods=["POST"])
    def api_peak_reset():
        month = request.json.get("month") if request.json else None
        if _peak:
            _peak.reset_monthly(month)
        return jsonify({"ok": True, "month": month or datetime.now().strftime("%Y-%m")})

    def _hm_diff(a: str, b: str) -> int:
        """Difference in minutes between two HH:MM strings."""
        try:
            ah, am = map(int, a.split(":"))
            bh, bm = map(int, b.split(":"))
            return (ah*60+am) - (bh*60+bm)
        except:
            return 999

    def _save_demand_snapshot(circles_now, hist_file):
        import os, json as _json
        from datetime import datetime
        now = datetime.now()
        entry = {
            "date": now.strftime("%Y-%m-%d"),
            "time": now.strftime("%H:%M"),
            "circles": circles_now,
        }
        hist = []
        if os.path.exists(hist_file):
            try:
                hist = _json.load(open(hist_file))
            except:
                pass
        hist.append(entry)
        # Keep last 2 days
        cutoff = (now - __import__("datetime").timedelta(days=2)).strftime("%Y-%m-%d")
        hist = [h for h in hist if h.get("date","") >= cutoff]
        _json.dump(hist, open(hist_file,"w"), indent=2)

    log.info("All API routes registered")

    @app.route("/api/email/discover", methods=["POST"])
    def api_email_discover():
        """Discover SMTP servers reachable from this machine."""
        import socket
        candidates = [
            # Internal relay candidates (common on corporate networks)
            ("localhost",       25),
            ("localhost",       587),
            ("127.0.0.1",       25),
            ("10.79.176.1",     25),   # likely gateway
            ("10.79.176.1",     587),
            ("10.79.176.254",   25),
            ("mail",            25),
            ("smtp",            25),
            ("mailrelay",       25),
            ("relay",           25),
            # External (likely blocked)
            ("smtp.gmail.com",  587),
            ("smtp.gmail.com",  465),
        ]
        results = []
        for host, port in candidates:
            try:
                s = socket.create_connection((host, port), timeout=3)
                banner = ""
                try:
                    s.settimeout(2)
                    banner = s.recv(256).decode("utf-8","ignore").strip()[:80]
                except Exception:
                    pass
                s.close()
                results.append({"host": host, "port": port,
                                 "ok": True, "banner": banner})
            except Exception:
                results.append({"host": host, "port": port,
                                 "ok": False, "banner": ""})
        reachable = [r for r in results if r["ok"]]
        return jsonify({"results": results, "reachable": reachable})

    # ── GSS Summary ──────────────────────────────────────────
    @app.route("/api/shutdown", methods=["GET","POST"])
    def api_shutdown():
        """Graceful shutdown — closes Flask sockets cleanly before process exits."""
        import threading, os as _os
        def _do_shutdown():
            import time as _t; _t.sleep(0.5)
            _os._exit(0)
        threading.Thread(target=_do_shutdown, daemon=True).start()
        return "Shutting down...", 200

    @app.route("/api/gss-summary")
    def api_gss_summary():
        """Per-GSS aggregated data: load, voltage status, alert counts, feeders."""
        data   = _scraper.last_data
        vn     = _cfg.get("voltage.vn_kv", 33.0)
        ov_lim = vn * (1 + _cfg.get("voltage.ov_pct", 6.0) / 100)
        uv_lim = vn * (1 - _cfg.get("voltage.uv_pct", 9.0) / 100)
        ol_thr = _cfg.get("voltage.load_threshold_pct", 100.0) / 100
        off_thr= _cfg.get("voltage.feeder_off_threshold_a", 1.0)
        meta   = getattr(_scraper, "_meta_lookup", {})

        # Active alert counts from DB
        active_alerts = _store.all(active_only=True, limit=1000)
        gss_alerts: dict = {}
        for a in active_alerts:
            gss = (a.get("Gss") or "").strip()
            if not gss:
                continue
            if gss not in gss_alerts:
                gss_alerts[gss] = {"OL":0,"FEEDER_OFF":0,"LOAD_DIVERTED":0,"OV":0,"UV":0}
            t = a.get("type","")
            if t in gss_alerts[gss]:
                gss_alerts[gss][t] += 1

        # Aggregate live data by GSS
        gss_data: dict = {}
        for r in data:
            ac  = r.get("AssetCode","")
            fm  = _fm.lookup(ac)
            gss = (fm.get("GssName","") if fm else "") or r.get("Gss","") or \
                  meta.get(ac,{}).get("Gss","")
            if not gss:
                continue

            vr = float(r.get("Vr") or 0)
            vy = float(r.get("Vy") or 0)
            vb = float(r.get("Vb") or 0)
            ir = float(r.get("Ir") or 0)
            iy = float(r.get("Iy") or 0)
            ib = float(r.get("Ib") or 0)
            vavg = (vr+vy+vb)/3
            imax = max(ir,iy,ib)
            raw_ap2 = float(r.get("ActivePower") or 0)
            raw_sp2 = float(r.get("ApparentPower") or 0)
            _fm_g  = _fm.lookup((r.get("AssetCode") or "").strip()) if '_fm_g' not in dir() else _fm.lookup((r.get("AssetCode") or "").strip())
            _sign2 = (_fm_g.get("PowerSign","") if _fm_g else "").strip().lower()
            if _sign2 == "-":
                ap, sp = -raw_ap2, -raw_sp2
            elif _sign2 == "solar":
                ap, sp = raw_ap2, raw_sp2
            else:
                ap, sp = abs(raw_ap2), abs(raw_sp2)
            fn   = (fm.get("FeederName","") if fm else "") or r.get("Feeder","") or ac
            is_bc= bool(fm.get("IsBusCoupler") if fm else False) or \
                   ("BUS" in fn.upper() and "COUPL" in fn.upper())

            if gss not in gss_data:
                circle   = (fm.get("CircleName","") if fm else "") or r.get("Circle","")
                division = (fm.get("DivisionName","") if fm else "") or r.get("Division","")
                gss_data[gss] = {
                    "Gss": gss, "Circle": circle, "Division": division,
                    "MW_now": 0, "MVA_now": 0,
                    "feeders": [], "feeder_count": 0, "bc_count": 0,
                    "v_status": "OK",  # worst: OV > UV > OK
                }

            g = gss_data[gss]
            g["MW_now"]  += ap
            g["MVA_now"] += sp
            g["feeder_count"] += 1
            if is_bc:
                g["bc_count"] += 1

            # Voltage status — same logic as /api/live (exclude BC, unbalance check)
            v_unbal_gs = (max(abs(vr-vavg), abs(vy-vavg), abs(vb-vavg)) / vavg * 100) if vavg > 1 else 0
            if vavg > ov_lim:
                vs = "OV"
            elif vavg < uv_lim and vavg > 1:
                if is_bc:
                    vs = "OK"
                elif v_unbal_gs > 25:
                    vs = "PT"
                elif v_unbal_gs < 10:
                    vs = "UV"
                else:
                    vs = "OK"
            else:
                vs = "OK"
            if vs == "OV" or (vs == "UV" and g["v_status"] != "OV"):
                g["v_status"] = vs

            # Feeder detail
            rating  = float(fm.get("FeederRating",0) if fm else 0)
            load_pct= round(imax/rating*100,1) if rating > 0 else None
            is_solar_feeder = bool((fm or {}).get("IsSolarPlant"))
            if is_bc:
                i_status = "OK" if imax < off_thr else "DIV"
            elif is_solar_feeder:
                # Same deferral as main /api/live — solar uses wait-time
                # confirmation, not a raw current check. Show OFF only if
                # the engine has actually confirmed it in DB.
                if _store.is_active(f"{ac}_FEEDER_OFF"):
                    i_status = "OFF"
                elif rating > 0 and imax/rating > ol_thr:
                    i_status = "OL"
                else:
                    i_status = "OK"
            elif imax < off_thr:
                i_status = "OFF"
            elif rating > 0 and imax/rating > ol_thr:
                i_status = "OL"
            else:
                i_status = "OK"

            g["feeders"].append({
                "AssetCode": ac, "Feeder": fn, "FeederType": fm.get("FeederType","") if fm else "",
                "IsBusCoupler": is_bc, "FeederRating": rating,
                "Vr":round(vr,3),"Vy":round(vy,3),"Vb":round(vb,3),
                "Vavg":round(vavg,3),
                "Ir":round(ir,2),"Iy":round(iy,2),"Ib":round(ib,2),"Imax":round(imax,2),
                "ActivePower":round(ap,3),"ApparentPower":round(sp,3),
                "LoadPct": load_pct, "V_status": vs, "I_status": i_status,
            })

        # Round totals + attach alert counts + per-GSS peaks
        result = []
        gss_peaks = _peak.get_gss_daily() if _peak else {}

        for gss, g in gss_data.items():
            g["MW_now"]  = round(g["MW_now"],  3)
            g["MVA_now"] = round(g["MVA_now"], 3)
            alerts = gss_alerts.get(gss, {})
            g["alert_ol"]  = alerts.get("OL", 0)
            g["alert_off"] = alerts.get("FEEDER_OFF", 0)
            g["alert_div"] = alerts.get("LOAD_DIVERTED", 0)
            g["alert_ov"]  = alerts.get("OV", 0)
            g["alert_uv"]  = alerts.get("UV", 0)
            # Count non-BC feeders in UV/OV (for threshold logic in frontend)
            g["uv_feeder_count"] = sum(1 for f in g["feeders"]
                                       if not f.get("IsBusCoupler") and f.get("V_status")=="UV")
            g["ov_feeder_count"] = sum(1 for f in g["feeders"]
                                       if not f.get("IsBusCoupler") and f.get("V_status")=="OV")

            # Active non-BC feeders (carrying current > off_thr)
            g["active_feeder_count"] = sum(1 for f in g["feeders"]
                                           if not f.get("IsBusCoupler")
                                           and f.get("Imax",0) > off_thr)

            # Real-time GSS voltage: average Vavg of healthy meters
            # Healthy = non-BC + carrying current + voltage unbalance < 10%
            healthy_vavgs = []
            for f in g["feeders"]:
                if f.get("IsBusCoupler"):
                    continue
                if f.get("Imax",0) < off_thr:
                    continue  # feeder off
                vr_ = f.get("Vr",0); vy_ = f.get("Vy",0); vb_ = f.get("Vb",0)
                v_avg_ = (vr_ + vy_ + vb_) / 3
                if v_avg_ < 1:
                    continue  # no voltage
                # Voltage unbalance % = max(|Vx - Vavg|) / Vavg * 100
                v_unbal = max(abs(vr_-v_avg_), abs(vy_-v_avg_), abs(vb_-v_avg_)) / v_avg_ * 100
                if v_unbal < 10:
                    healthy_vavgs.append(v_avg_)
            g["gss_voltage"] = round(sum(healthy_vavgs)/len(healthy_vavgs), 3) if healthy_vavgs else None
            g["healthy_meter_count"] = len(healthy_vavgs)
            # Per-GSS daily peak
            gp = gss_peaks.get(gss, {})
            g["peak_mw"]   = round(gp.get("peak_mw",  g["MW_now"]), 3)
            g["peak_mva"]  = round(gp.get("peak_mva", g["MVA_now"]), 3)
            g["peak_time"] = gp.get("peak_time", "")
            g["feeders"].sort(key=lambda f: (f["IsBusCoupler"], f["Feeder"]))
            result.append(g)

        # Sort: by Circle then GSS name
        result.sort(key=lambda g: (g.get("Circle",""), g.get("Gss","")))
        return jsonify({"gss": result, "count": len(result),
                        "fetched_at": _scraper.last_fetch_time})
